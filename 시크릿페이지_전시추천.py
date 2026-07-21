"""
LF몰 제휴 시크릿 페이지 큐레이션 대시보드
"""

import streamlit as st
import pandas as pd
import numpy as np
import json
import io
from datetime import datetime

# ─────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="LF몰 시크릿 큐레이션",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: 'Pretendard', 'Apple SD Gothic Neo', 'Noto Sans KR', sans-serif;
}
section[data-testid="stSidebar"] { background: #1a2a40; }
section[data-testid="stSidebar"] * { color: #FFFFFF !important; }
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div { color: #FFFFFF !important; }
section[data-testid="stSidebar"] label {
    color: #D0DFF0 !important; font-size: 0.82rem; font-weight: 600;
}
section[data-testid="stSidebar"] .stMarkdown { color: #FFFFFF !important; }
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 { color: #FFFFFF !important; font-weight: 700; }
section[data-testid="stSidebar"] .stCaption { color: #DCE8F5 !important; }
section[data-testid="stSidebar"] .stCaption * { color: #DCE8F5 !important; }
section[data-testid="stSidebar"] small { color: #DCE8F5 !important; }
section[data-testid="stSidebar"] .badge-ok   { background: #1a5c30; color: #7FFFA0 !important; border-radius:4px; padding:2px 8px; font-size:.72rem; font-weight:700; }
section[data-testid="stSidebar"] .badge-wait { background: #5c4a00; color: #FFE066 !important; border-radius:4px; padding:2px 8px; font-size:.72rem; font-weight:700; }

/* 입력 위젯: 위의 * 규칙이 글자만 흰색으로 만들어 배경과 겹치는 것을 방지.
   배경·글자색을 함께 명시해 대비를 확보한다. */
section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] textarea,
section[data-testid="stSidebar"] [data-baseweb="input"],
section[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: #FFFFFF !important;
    color: #16243A !important;
    border: 1px solid #4A6488 !important;
    border-radius: 6px !important;
}
section[data-testid="stSidebar"] input::placeholder,
section[data-testid="stSidebar"] textarea::placeholder {
    color: #7A8AA0 !important;
    opacity: 1;
}
section[data-testid="stSidebar"] [data-baseweb="select"] svg { fill: #16243A !important; }

/* 파일 업로더 드롭존 */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
    background: #24374F !important;
    border: 1px dashed #6E88AB !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] * {
    color: #E8F0FA !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
    background: #FFFFFF !important;
    color: #16243A !important;
    border: none !important;
    font-weight: 600;
}
/* 업로드된 파일명 */
section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] * {
    color: #E8F0FA !important;
}

/* 기간 설정 안내(분석 ↔ 전년)의 주차코드 칩 */
section[data-testid="stSidebar"] code,
section[data-testid="stSidebar"] .stCaption code {
    background: #33506F !important;
    color: #FFDCA8 !important;
    padding: 1px 6px;
    border-radius: 4px;
    font-weight: 700;
}

/* 셀렉트박스에 표시되는 선택값 */
section[data-testid="stSidebar"] [data-baseweb="select"] div[value],
section[data-testid="stSidebar"] [data-baseweb="select"] span {
    color: #16243A !important;
}

/* 드롭다운 목록은 사이드바 밖(문서 최상단)에 그려지므로 따로 지정한다.
   사이드바 스코프 규칙이 닿지 않아 흰 배경에 흰 글씨가 되기 쉽다. */
div[data-baseweb="popover"] li,
div[data-baseweb="popover"] [role="option"] {
    background: #FFFFFF !important;
    color: #16243A !important;
}
div[data-baseweb="popover"] li:hover,
div[data-baseweb="popover"] [role="option"]:hover {
    background: #E8EFF8 !important;
    color: #16243A !important;
}
.main .block-container { background: #F5F6F8; padding-top: 1.5rem; }

.header-banner {
    background: linear-gradient(135deg, #1E3A5F 0%, #0F1B2D 70%);
    border-radius: 12px; padding: 1.6rem 2rem; margin-bottom: 1.5rem;
    display: flex; justify-content: space-between; align-items: center;
}
.header-title  { color: #FFF; font-size: 1.4rem; font-weight: 700; letter-spacing: -0.02em; }
.header-sub    { color: #7BA3CC; font-size: 0.82rem; margin-top: .25rem; }
.header-badge  {
    background: rgba(255,255,255,.08); border: 1px solid rgba(255,255,255,.15);
    border-radius: 20px; padding: .3rem .9rem; color: #A8D4FF; font-size: .78rem;
}

.kpi-row { display: flex; gap: 12px; margin-bottom: 1.5rem; flex-wrap: wrap; }
.kpi-card {
    flex: 1; min-width: 130px;
    background: #FFF; border-radius: 10px; padding: 1rem 1.2rem;
    border-left: 3px solid #1E3A5F; box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.kpi-label { color: #6B7A8D; font-size: .7rem; font-weight: 600; letter-spacing: .06em; text-transform: uppercase; }
.kpi-value { color: #1E3A5F; font-size: 1.6rem; font-weight: 700; margin-top: .2rem; line-height: 1; }
.kpi-sub   { color: #9BA8B5; font-size: .7rem; margin-top: .25rem; }

.sec-hdr {
    background: #FFF; border-radius: 10px 10px 0 0;
    padding: .85rem 1.2rem; border-bottom: 2px solid #F0F2F5;
    display: flex; align-items: center; gap: .6rem;
}
.sec-num {
    background: #1E3A5F; color: #FFF; border-radius: 6px;
    width: 24px; height: 24px; display: inline-flex;
    align-items: center; justify-content: center;
    font-size: .72rem; font-weight: 700; flex-shrink: 0;
}
.sec-title { color: #1E3A5F; font-size: .88rem; font-weight: 700; }
.sec-desc  { color: #8A96A3; font-size: .72rem; margin-left: auto; }

.tbl-wrap {
    background: #FFF; border-radius: 0 0 10px 10px;
    padding: 1rem; box-shadow: 0 1px 4px rgba(0,0,0,.06); margin-bottom: 1.5rem;
}

.badge-ok   { background: #E6F4EA; color: #2D7D46; border-radius: 4px; padding: 2px 8px; font-size: .7rem; font-weight: 600; }
.badge-wait { background: #FFF3CD; color: #8A6914; border-radius: 4px; padding: 2px 8px; font-size: .7rem; font-weight: 600; }

.w-chip         { display:inline-block; border-radius:12px; padding:.15rem .6rem; font-size:.68rem; font-weight:600; margin-right:4px; }
.w-chip.trend   { background:#FFF0F6; color:#D63E8A; }
.w-chip.vol     { background:#F0FFF4; color:#2D7D46; }
.w-chip.growth  { background:#FFF8E1; color:#B45309; }
.w-chip.wow     { background:#EEF2FF; color:#4338CA; }

.info-box { background:#EEF4FF; border-radius:8px; padding:.8rem 1rem; color:#2255A4; font-size:.78rem; margin:.5rem 0; }
.warn-box { background:#FFF8E1; border-radius:8px; padding:.8rem 1rem; color:#7B5A00; font-size:.78rem; margin:.5rem 0; }

.stTabs [data-baseweb="tab-list"] {
    gap:4px; background:#F0F2F5; border-radius:8px; padding:4px;
}
.stTabs [data-baseweb="tab"] {
    border-radius:6px; font-size:.82rem; font-weight:600; color:#6B7A8D; padding:6px 14px;
}
.stTabs [aria-selected="true"] { background:#FFF !important; color:#1E3A5F !important; }

.stDownloadButton > button {
    background:#1E3A5F !important; color:white !important;
    border-radius:8px !important; border:none !important;
    font-weight:600 !important; font-size:.82rem !important;
}
.stDownloadButton > button:hover { background:#2A5080 !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def _to_float(v):
    """콤마 포함 문자열, None, NaN 모두 float으로 안전 변환"""
    try:
        if pd.isna(v): return None
    except Exception:
        pass
    try:
        return float(str(v).replace(',', ''))
    except Exception:
        return None

def fmt_amt(v):
    v = _to_float(v)
    if v is None or v == 0: return "0"
    if abs(v) >= 1e8: return f"{v/1e8:.1f}억"
    if abs(v) >= 1e4: return f"{v/1e4:.0f}만"
    return f"{v:,.0f}"

def fmt_num(v):
    v = _to_float(v)
    if v is None: return "-"
    return f"{int(v):,}"

def fmt_pct(v):
    v = _to_float(v)
    if v is None: return "-"
    sign = "+" if v > 0 else ""
    return f"{sign}{v:.1f}%"

def fmt_int_signed(v):
    v = _to_float(v)
    if v is None: return "-"
    sign = "+" if v > 0 else ""
    return f"{sign}{int(v):,}"

def mall_week_to_code(week_str):
    """'2025년 06월 3주차' → '25_6_3'"""
    try:
        s = week_str.replace("년","_").replace("월","_").replace("주차","").strip()
        parts = [p.strip() for p in s.split("_") if p.strip()]
        y = parts[0][-2:]
        m = str(int(parts[1]))
        w = parts[2]
        return f"{y}_{m}_{w}"
    except:
        return week_str

def code_to_mall_week(code, year=2025):
    """'25_6_3' → '2025년 06월 3주차'"""
    try:
        parts = code.split("_")
        y = 2000 + int(parts[0])
        m = int(parts[1])
        w = parts[2]
        return f"{y}년 {m:02d}월 {w}주차"
    except:
        return code

def get_prev_week_code(code):
    """'26_6_3' → '25_6_3' (전년 동주차)"""
    try:
        parts = code.split("_")
        return f"{int(parts[0])-1}_{parts[1]}_{parts[2]}"
    except:
        return None

def get_next_week_code(code, available_codes=None):
    """
    '25_7_4' → '25_8_1' (한 주 뒤 주차)

    한 달이 4주인지 5주인지는 달마다 달라 산술만으로 정할 수 없다.
    데이터에 실제 존재하는 주차 목록(available_codes)이 있으면 그걸 기준으로
    다음 주차가 같은 달에 있는지 판단하고, 없으면 다음 달 1주차로 넘긴다.
    목록이 없을 때는 5주차까지 있을 수 있다고 보고 판단한다.
    """
    try:
        y, m, w = (int(x) for x in code.split("_"))
    except Exception:
        return None

    nxt = f"{y}_{m}_{w + 1}"
    if available_codes is not None:
        if nxt in available_codes:
            return nxt
    elif w < 5:
        return nxt
    return f"{y}_{m + 1}_1" if m < 12 else f"{y + 1}_1_1"

def get_prev_adjacent_week_code(code, available_codes=None):
    """
    '25_8_1' → '25_7_4' (바로 앞 주차)

    get_next_week_code 의 반대 방향. 1주차에서 앞 달로 넘어갈 때
    그 달이 4주까지인지 5주까지인지 데이터로 확인한다.
    """
    try:
        y, m, w = (int(x) for x in code.split("_"))
    except Exception:
        return None

    if w > 1:
        return f"{y}_{m}_{w - 1}"

    py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
    if available_codes:
        for cand in range(5, 0, -1):
            c = f"{py}_{pm}_{cand}"
            if c in available_codes:
                return c
    return f"{py}_{pm}_4"

def build_week_map(df_pivot):
    """피벗 E열↔G열: {날짜str: 주차코드}"""
    wmap = {}
    for _, row in df_pivot.iterrows():
        dval = row.iloc[4]
        wcode = row.iloc[6]
        if pd.notna(dval) and pd.notna(wcode):
            try:
                d = pd.to_datetime(dval)
                wmap[d.strftime('%Y-%m-%d')] = str(wcode).strip()
            except:
                pass
    return wmap

def build_affiliate_map(df_pivot):
    """피벗 B열↔C열: {코드: 제휴사명}"""
    amap = {}
    for _, row in df_pivot.iterrows():
        code = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        name = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
        if code and name and code not in ('<제휴사 구분>', 'nan'):
            amap[code] = name
    return amap

def get_vat_col(df):
    """거래액_VAT제외 컬럼 반환. 없으면 거래액/1.1 계산 후 반환"""
    if '거래액_VAT제외' in df.columns:
        return pd.to_numeric(df['거래액_VAT제외'], errors='coerce')
    return pd.to_numeric(df['거래액'], errors='coerce') / 1.1

def tag_week(df, week_map):
    """정산일시일 → 주차코드 컬럼 추가"""
    df = df.copy()
    df['_date'] = pd.to_datetime(df['정산일시일']).dt.strftime('%Y-%m-%d')
    df['_week'] = df['_date'].map(week_map)
    return df

def build_dashboard_html(sel_week, prev_week, kpis, sections):
    """
    화면 내용을 그대로 담은 단일 HTML 파일을 만든다.

    첨부·공유를 전제로 하므로 외부 CSS/JS/폰트를 참조하지 않고
    스타일을 전부 인라인으로 넣는다. 인터넷 없이도 열린다.
    sections: [(번호, 제목, 설명, DataFrame), ...]
    """
    def esc(v):
        return (str(v).replace('&', '&amp;').replace('<', '&lt;')
                .replace('>', '&gt;').replace('"', '&quot;'))

    def table_html(df):
        if df is None or df.empty:
            return '<p class="empty">해당 주차 데이터 없음.</p>'
        head = ''.join(f'<th>{esc(c)}</th>' for c in df.columns)
        rows = []
        for row in df.itertuples(index=False):
            cells = ''.join(f'<td>{esc(v)}</td>' for v in row)
            rows.append(f'<tr>{cells}</tr>')
        return (f'<table><thead><tr>{head}</tr></thead>'
                f'<tbody>{"".join(rows)}</tbody></table>')

    kpi_html = ''.join(
        f'<div class="kpi"><div class="kpi-label">{esc(label)}</div>'
        f'<div class="kpi-value" style="color:{color}">{esc(value)}</div>'
        f'<div class="kpi-sub">{esc(sub)}</div></div>'
        for label, value, sub, color in kpis
    )

    body = []
    for num, title, desc, df in sections:
        body.append(
            f'<section><h2><span class="num">{esc(num)}</span>{esc(title)}'
            f'<span class="desc">{esc(desc)}</span></h2>{table_html(df)}</section>'
        )

    generated = datetime.now().strftime('%Y-%m-%d %H:%M')
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>LF몰 시크릿 페이지 큐레이션 {esc(sel_week)}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin:0; padding:24px; background:#F5F6F8; color:#16243A;
         font-family:'Pretendard','Apple SD Gothic Neo','Noto Sans KR',sans-serif; }}
  .wrap {{ max-width:1200px; margin:0 auto; }}
  .banner {{ background:linear-gradient(135deg,#1E3A5F 0%,#0F1B2D 70%);
             border-radius:12px; padding:1.2rem 1.5rem; display:flex;
             justify-content:space-between; align-items:center; gap:1rem; flex-wrap:wrap; }}
  .banner h1 {{ color:#FFF; font-size:1.3rem; margin:0; font-weight:700; }}
  .banner .sub {{ color:#7BA3CC; font-size:.82rem; margin-top:.3rem; }}
  .badge {{ background:rgba(255,255,255,.08); border:1px solid rgba(255,255,255,.15);
            border-radius:20px; padding:.3rem .9rem; color:#A8D4FF; font-size:.78rem; }}
  .meta {{ color:#6B7A8D; font-size:.76rem; margin:.9rem 0 1.2rem; }}
  .kpi-row {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr));
              gap:12px; margin-bottom:1.5rem; }}
  .kpi {{ background:#FFF; border-radius:10px; padding:1rem 1.2rem;
          border-left:4px solid #1E3A5F; }}
  .kpi-label {{ color:#6B7A8D; font-size:.7rem; font-weight:600;
                letter-spacing:.06em; text-transform:uppercase; }}
  .kpi-value {{ color:#1E3A5F; font-size:1.6rem; font-weight:700; margin-top:.2rem; }}
  .kpi-sub {{ color:#9BA8B5; font-size:.7rem; margin-top:.25rem; }}
  section {{ background:#FFF; border-radius:10px; padding:1rem 1.2rem; margin-bottom:1.2rem; }}
  h2 {{ font-size:.9rem; color:#1E3A5F; display:flex; align-items:center;
        gap:.5rem; margin:0 0 .8rem; }}
  .num {{ background:#1E3A5F; color:#FFF; border-radius:6px;
          padding:2px 8px; font-size:.78rem; }}
  .desc {{ color:#8A96A3; font-size:.72rem; font-weight:400; margin-left:auto; }}
  .tbl {{ overflow-x:auto; }}
  table {{ border-collapse:collapse; width:100%; font-size:.8rem; }}
  th {{ background:#F0F2F5; color:#1E3A5F; text-align:left;
        padding:8px 10px; border-bottom:2px solid #DDE2E8; white-space:nowrap; }}
  td {{ padding:7px 10px; border-bottom:1px solid #EEF0F3; white-space:nowrap; }}
  tbody tr:nth-child(even) {{ background:#FAFBFC; }}
  .empty {{ color:#8A6914; background:#FFF8E1; border-radius:8px;
            padding:.8rem 1rem; font-size:.8rem; }}
  footer {{ color:#9BA8B5; font-size:.72rem; text-align:center; margin-top:1.5rem; }}
  @media print {{ body {{ background:#FFF; padding:0; }} section {{ break-inside:avoid; }} }}
</style>
</head>
<body>
<div class="wrap">
  <div class="banner">
    <div>
      <h1>🏷️ LF몰 시크릿 페이지 큐레이션 대시보드</h1>
      <div class="sub">전년 트렌드 × 제휴 실적 × 신장률 → 게시 브랜드 &amp; 상품 선정</div>
    </div>
    <div class="badge">내부 전용 · CONFIDENTIAL</div>
  </div>
  <div class="meta">분석 주차 <b>{esc(sel_week)}</b> ↔ 전년 비교 <b>{esc(prev_week)}</b>
    · 생성 {esc(generated)}</div>
  <div class="kpi-row">{kpi_html}</div>
  {''.join(body)}
  <footer>LF몰 시크릿 페이지 큐레이션 대시보드 · 내부 전용</footer>
</div>
</body>
</html>"""


def current_report_html():
    """session_state 의 분석 결과로 대시보드 HTML 을 만든다."""
    w = st.session_state.sel_week
    pw = st.session_state.sel_prev_mall_week
    prev_code = mall_week_to_code(pw) if pw else '-'
    n1, n2, n3, n4 = (len(st.session_state[f'df{i}']) for i in range(1, 5))
    n_wow = len(st.session_state.df_wow)
    base_prev = st.session_state.base_prev_mall

    fmt_cols = {'거래액': fmt_amt, '고객수': fmt_num, '객단가': fmt_amt,
                '전주비(%)': fmt_pct, '전년비(%)': fmt_pct,
                '고객수증감': fmt_int_signed, '고객수증감률(%)': fmt_pct,
                '거래액전년비(%)': fmt_pct, '고객수효과': fmt_amt, '객단가효과': fmt_amt}

    def formatted(df):
        if df is None or df.empty:
            return df
        out = df.copy()
        for col, fn in fmt_cols.items():
            if col in out.columns:
                out[col] = out[col].apply(fn)
        return out

    return build_dashboard_html(
        sel_week=w,
        prev_week=prev_code,
        kpis=[
            ("전년 트렌드 TOP",  n1,    f"몰전체 · {prev_code}", "#1E3A5F"),
            ("전주비 신장 TOP",  n_wow, "몰전체 · 전주 대비",     "#1E3A5F"),
            ("제휴 볼륨 TOP",    n2,    f"제휴 · {w}",           "#1E3A5F"),
            ("전년비 신장 TOP",  n3,    "전년 동주차 대비",       "#1E3A5F"),
            ("최종 선정",        n4,    "가중치 종합",            "#D63E8A"),
        ],
        sections=[
            ("1", "전년 트렌드 TOP 20", f"몰전체 {prev_code}", formatted(st.session_state.df1)),
            ("2", "전년 전주비 신장 TOP 30",
             f"몰전체 {mall_week_to_code(base_prev) if base_prev else '-'} → {prev_code}",
             formatted(st.session_state.df_wow)),
            ("3", "제휴 볼륨 TOP 20",   f"제휴 {w}",           formatted(st.session_state.df2)),
            ("4", "전년비 신장 TOP 20", "전년 동주차 대비",     formatted(st.session_state.df3)),
            ("5", "최종 선정 30개",     "가중치 종합 점수",     formatted(st.session_state.df4)),
            ("6", "브랜드별 상품코드 TOP 10", "선택 브랜드 기준",
             formatted(st.session_state.df_prod)),
        ],
    )


def df_to_excel_bytes(df, sheet_name='Sheet1'):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        hdr_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_align = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
        for col_cells in ws.columns:
            max_w = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_w + 2, 28)
    buf.seek(0)
    return buf.getvalue()

def dfs_to_json_bytes(data_dict):
    out = {}
    for key, df in data_dict.items():
        if df is not None:
            dc = df.copy()
            for col in dc.select_dtypes(include=['datetime64']).columns:
                dc[col] = dc[col].astype(str)
            out[key] = dc.to_dict(orient='records')
        else:
            out[key] = []
    return json.dumps(out, ensure_ascii=False, indent=2).encode('utf-8')


# ─────────────────────────────────────────────
# 로딩 함수 (캐시)
# ─────────────────────────────────────────────
def needs_parse(widget_key, uploaded):
    """
    같은 파일을 재실행마다 다시 파싱하지 않도록 잠근다.

    Streamlit은 위젯을 건드릴 때마다 스크립트를 처음부터 다시 돌린다.
    그대로 두면 재실행마다 원본을 재파싱하면서, 새로 만든 DataFrame과
    session_state 의 기존 DataFrame 이 동시에 메모리에 올라간다.
    파일이 바뀌었을 때만 True 를 돌려준다.
    """
    return st.session_state.get(f'_sig_{widget_key}') != (uploaded.name, uploaded.size)


def mark_parsed(widget_key, uploaded):
    """
    파싱에 성공했을 때만 서명을 남긴다.

    시도 시점에 기록하면, 로드가 실패하거나 도중에 앱이 재시작됐을 때
    서명만 남아 다음 재실행에서 '이미 처리한 파일'로 건너뛰게 된다.
    그러면 배지도 에러도 없이 파일이 멈춘 것처럼 보인다.
    """
    st.session_state[f'_sig_{widget_key}'] = (uploaded.name, uploaded.size)


def clear_parsed(widget_key):
    st.session_state.pop(f'_sig_{widget_key}', None)


def shrink_dtypes(df):
    """
    메모리 절감용 dtype 축소.
    - 정수/실수는 표현 가능한 최소 폭으로 downcast
    - 반복이 많은 문자열 컬럼(브랜드/카테고리 등)은 category 로 변환
    고유값이 절반을 넘는 컬럼은 category 가 오히려 손해라 건드리지 않는다.
    """
    for col in df.columns:
        s = df[col]
        if pd.api.types.is_integer_dtype(s):
            df[col] = pd.to_numeric(s, downcast='integer')
        elif pd.api.types.is_float_dtype(s):
            df[col] = pd.to_numeric(s, downcast='float')
        # pandas 3.x 는 문자열을 object 가 아니라 str dtype 으로 읽으므로 둘 다 본다.
        elif pd.api.types.is_object_dtype(s) or pd.api.types.is_string_dtype(s):
            # 날짜 컬럼은 뒤에서 pd.to_datetime 으로 파싱되므로 category 로 바꾸지 않는다.
            if any(k in str(col) for k in ('일시', '일자', '날짜', 'date', 'Date')):
                continue
            if pd.api.types.infer_dtype(s, skipna=True) != 'string':
                continue
            # 숫자가 문자열로 들어온 컬럼 처리 (엑셀 export 에서 흔함).
            # 문자열로 두면 sum() 이 덧셈이 아니라 문자열 이어붙이기가 되어
            # 값이 조용히 틀린다("10"+"20" -> "1020"). 그래서 숫자로 바꾼다.
            # 단 상품코드처럼 앞자리 0 이 의미를 갖는 값은 문자열로 남긴다.
            nonnull = s.dropna().astype(str)
            if len(nonnull):
                stripped = nonnull.str.replace(',', '', regex=False).str.strip()
                converted = pd.to_numeric(stripped, errors='coerce')
                if converted.notna().mean() >= 0.95:
                    has_leading_zero = (
                        stripped.str.len().gt(1) & stripped.str.startswith('0')
                    ).any()
                    if not has_leading_zero:
                        num = pd.to_numeric(
                            s.astype(str).str.replace(',', '', regex=False).str.strip(),
                            errors='coerce',
                        )
                        kind = 'integer' if (num.dropna() % 1 == 0).all() else 'float'
                        df[col] = pd.to_numeric(num, downcast=kind)
                    continue
            n = len(s)
            if n and s.nunique(dropna=False) / n < 0.5:
                df[col] = s.astype('category')
    return df


def load_mall_csv(file_bytes):
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), encoding='utf-16', sep='\t')
        df.columns = [c.strip() for c in df.columns]
        if '거래액' in df.columns:
            df['거래액'] = pd.to_numeric(
                df['거래액'].astype(str).str.replace(',', ''), errors='coerce'
            )
        # 주차코드 컬럼 추가 (피벗 없이 자체 변환)
        if '결제_주차' in df.columns:
            df['_week_code'] = df['결제_주차'].apply(mall_week_to_code)
        return shrink_dtypes(df), None
    except Exception as e:
        return None, str(e)

def load_affiliate_excel(file_bytes, sheet_hint='26'):
    """
    인증거래액 시트 + 피벗 시트 로드.
    sheet_hint: '26' → '인증거래액_26', '25' → '인증거래액_4-6월' 등 자동 탐색

    메모리 주의: 이 앱은 2.7GB 안에서 돌아야 한다.
    - 엔진은 calamine 고정. openpyxl은 셀마다 파이썬 객체를 만들어
      같은 파일에 수 배의 메모리를 쓴다.
    - BytesIO는 생성할 때마다 바이트를 복사하므로 하나만 만들어
      seek(0)으로 재사용한다.
    """
    try:
        buf = io.BytesIO(file_bytes)

        with pd.ExcelFile(buf, engine='calamine') as xls:
            sheets = xls.sheet_names

            # 전년 파일은 분기별로 '인증거래액_1-3월/4-6월/7-9월/10-12월'처럼
            # 나뉘어 있을 수 있어, 이름에 '인증거래액'이 든 시트를 '전부' 읽어 병합한다.
            # (첫 시트만 읽으면 전년 동주차가 다른 분기라 신장률 TOP이 빈 표가 됨)
            auth_sheets = [s for s in sheets if '인증거래액' in s]
            if not auth_sheets:
                return None, None, f"'인증거래액' 시트를 찾을 수 없습니다. 시트목록: {sheets}"

            if len(auth_sheets) == 1:
                df_raw = shrink_dtypes(xls.parse(auth_sheets[0]))
            else:
                df_raw = pd.concat(
                    [shrink_dtypes(xls.parse(s)) for s in auth_sheets],
                    ignore_index=True,
                )

            df_piv = None
            if '피벗' in sheets:
                df_piv = xls.parse('피벗', header=None)

        return df_raw, df_piv, None
    except Exception as e:
        return None, None, str(e)


# ─────────────────────────────────────────────
# 분석 함수
# ─────────────────────────────────────────────

def analyze_mall_top20(df_mall, target_mall_week):
    """
    [분석1] 전년 몰전체 특정 주차 판매 TOP 20
    출력: 카테고리/브랜드/고객수/거래액
    (자사입점·정상이월은 몰전체 CSV에 없으므로 '-' 처리)
    """
    if df_mall is None or df_mall.empty:
        return pd.DataFrame()

    df = df_mall[df_mall['결제_주차'] == target_mall_week].copy()
    if df.empty:
        return pd.DataFrame()

    grp = df.groupby(['대카테고리명', 'ADMIN브랜드명'], as_index=False).agg(
        고객수=('주문고객수', 'sum'),
        거래액=('거래액', 'sum')
    ).sort_values('거래액', ascending=False).head(20).reset_index(drop=True)
    grp.insert(0, '순위', range(1, len(grp)+1))
    grp['자사/입점'] = '-'
    grp['정상/이월'] = '-'
    grp = grp[['순위','대카테고리명','ADMIN브랜드명','자사/입점','정상/이월','고객수','거래액']]
    grp.columns = ['순위','카테고리','브랜드','자사/입점','정상/이월','고객수','거래액']
    return grp


def analyze_mall_wow_growth(df_mall, base_mall_week, next_mall_week,
                            base_top_n=100, top_n=30):
    """
    [분석2] 전년 전주비 신장 TOP 30

    전년 분석주차(base)의 몰전체 거래액 TOP{base_top_n} 브랜드를 모수로 잡고,
    그 다음 주차(next)의 거래액 신장률이 높은 순으로 {top_n}개를 뽑는다.
    모수를 상위 브랜드로 한정하는 이유는, 거래액이 미미한 브랜드가
    작은 절대 증가만으로 신장률 상위를 차지하는 것을 막기 위해서다.

    표시하는 고객수·거래액은 다음 주차 기준이다(게시 대상 주차).
    자사/입점·정상/이월은 몰전체 CSV에 없어 '-' 로 채운다(분석1과 동일).
    """
    if df_mall is None or df_mall.empty or not base_mall_week or not next_mall_week:
        return pd.DataFrame()

    keys = ['대카테고리명', 'ADMIN브랜드명']

    def agg_week(week):
        d = df_mall[df_mall['결제_주차'] == week]
        if d.empty:
            return pd.DataFrame(columns=keys + ['고객수', '거래액'])
        return d.groupby(keys, as_index=False).agg(
            고객수=('주문고객수', 'sum'),
            거래액=('거래액', 'sum'),
        )

    base = agg_week(base_mall_week)
    if base.empty:
        return pd.DataFrame()

    base = base.sort_values('거래액', ascending=False).head(base_top_n)
    nxt = agg_week(next_mall_week)

    merged = base.merge(nxt, on=keys, how='left', suffixes=('_기준', ''))
    merged['고객수'] = merged['고객수'].fillna(0)
    merged['거래액'] = merged['거래액'].fillna(0)

    # 기준 주차 거래액이 0이면 신장률을 정의할 수 없으므로 제외한다.
    merged = merged[merged['거래액_기준'] > 0].copy()
    if merged.empty:
        return pd.DataFrame()

    merged['전주비(%)'] = (
        (merged['거래액'] - merged['거래액_기준']) / merged['거래액_기준'] * 100
    )

    merged = merged.sort_values('전주비(%)', ascending=False).head(top_n).reset_index(drop=True)
    merged.insert(0, '순위', range(1, len(merged) + 1))
    merged['자사/입점'] = '-'
    merged['정상/이월'] = '-'
    merged = merged[['순위'] + keys + ['자사/입점', '정상/이월', '고객수', '거래액', '전주비(%)']]
    merged.columns = ['순위', '카테고리', '브랜드', '자사/입점', '정상/이월',
                      '고객수', '거래액', '전주비(%)']
    return merged


def analyze_affiliate_top20_vol(df_raw, week_map, target_week_code):
    """
    [분석2] 제휴 실적 분석주차 판매 TOP 20 (당월인증 Y/N 모두 포함)
    출력: 카테고리/브랜드/정상이월/고객수/거래액/객단가
    """
    if df_raw is None or df_raw.empty or not week_map:
        return pd.DataFrame()

    df = tag_week(df_raw[df_raw['정산구분'] == '판매'], week_map)
    df = df[df['_week'] == target_week_code].copy()
    if df.empty:
        return pd.DataFrame()

    df['amt'] = get_vat_col(df)
    grp = df.groupby(['물리대카테', 'Admin브랜드명', '정상이월구분'], as_index=False).agg(
        고객수=('고객번호', 'nunique'),
        거래액=('amt', 'sum')
    )
    grp['객단가'] = (grp['거래액'] / grp['고객수']).round(0)
    grp = grp.sort_values('거래액', ascending=False).head(20).reset_index(drop=True)
    grp.insert(0, '순위', range(1, len(grp)+1))
    grp.columns = ['순위','카테고리','브랜드','정상/이월','고객수','거래액','객단가']
    return grp


# 신장률 TOP 튜닝 상수
GROWTH_MIN_CUST = 5        # (A) 당년·전년 고객수 둘 다 이 값 이상만 (소표본 폭주 방지)
GROWTH_MIN_AMT  = 300000   # 최소 당년 거래액(원, VAT제외)
GROWTH_REQUIRE_AMT_UP = True  # (C) 거래액도 전년비 플러스인 브랜드만(머릿수만 늘고 매출 빠진 브랜드 제외)

def analyze_affiliate_top20_growth(df_curr, df_prev, week_map_curr, week_map_prev,
                                    target_week_code):
    """
    [분석3] 전년 동주차 대비 '고객수 신장' TOP 20
    df_curr: 2026 raw / df_prev: 2025 raw · 전년 동주차: 26_6_3 → 25_6_3

    브랜드/주차 단위 고객수가 작아 거래액 %는 객단가 한 건에 크게 흔들린다.
    → (A) 최소 고객수 필터, (B) 고객수 증감을 주 랭킹(절대 증감분 우선, 율 병기),
       (C) 거래액은 가드레일 + 요인분해(고객수효과/객단가효과)로 병기.
    """
    if df_curr is None or df_curr.empty:
        return pd.DataFrame()

    prev_week_code = get_prev_week_code(target_week_code)

    # 당년 집계
    df_c = tag_week(df_curr[df_curr['정산구분'] == '판매'], week_map_curr)
    df_c = df_c[df_c['_week'] == target_week_code].copy()
    df_c['amt'] = get_vat_col(df_c)

    grp_c = df_c.groupby(['물리대카테', 'Admin브랜드명'], as_index=False).agg(
        고객수_당=('고객번호', 'nunique'),
        거래액_당=('amt', 'sum')
    )

    # 전년 집계 (2025 raw 있을 때) — 전년 고객수도 함께
    grp_p = pd.DataFrame(columns=['물리대카테', 'Admin브랜드명', '고객수_전', '거래액_전'])
    if df_prev is not None and not df_prev.empty and week_map_prev and prev_week_code:
        df_p = tag_week(df_prev[df_prev['정산구분'] == '판매'], week_map_prev)
        df_p = df_p[df_p['_week'] == prev_week_code].copy()
        df_p['amt'] = get_vat_col(df_p)
        if not df_p.empty:
            grp_p = df_p.groupby(['물리대카테', 'Admin브랜드명'], as_index=False).agg(
                고객수_전=('고객번호', 'nunique'),
                거래액_전=('amt', 'sum')
            )

    merged = grp_c.merge(grp_p, on=['물리대카테', 'Admin브랜드명'], how='left')

    # (A) 최소 표본: 당년·전년 고객수 둘 다 기준 이상 + 최소 거래액
    merged = merged[
        (merged['고객수_당'] >= GROWTH_MIN_CUST) &
        (merged['고객수_전'].fillna(0) >= GROWTH_MIN_CUST) &
        (merged['거래액_당'] >= GROWTH_MIN_AMT)
    ].copy()
    if merged.empty:
        return pd.DataFrame()

    # (B) 고객수 증감(주 지표) — 절대 + 율
    merged['고객수증감'] = merged['고객수_당'] - merged['고객수_전']
    merged['고객수증감률(%)'] = (merged['고객수증감'] / merged['고객수_전'] * 100)

    # 거래액 전년비(가드레일 표시용)
    merged['거래액전년비(%)'] = (merged['거래액_당'] - merged['거래액_전']) / merged['거래액_전'] * 100

    # (C) 요인분해: Δ거래액 = 고객수효과 + 객단가효과 (객단가_전 기준, 교차항은 객단가효과에 포함)
    aov_prev = merged['거래액_전'] / merged['고객수_전']
    merged['고객수효과'] = merged['고객수증감'] * aov_prev
    merged['객단가효과'] = (merged['거래액_당'] - merged['거래액_전']) - merged['고객수효과']

    # (C) 가드레일: 거래액도 전년비 플러스인 브랜드만
    if GROWTH_REQUIRE_AMT_UP:
        merged = merged[merged['거래액전년비(%)'] >= 0]
    if merged.empty:
        return pd.DataFrame()

    # (B) 랭킹: 절대 증감분 우선(규모 안정성), 동률은 증감률
    merged = merged.sort_values(['고객수증감', '고객수증감률(%)'],
                                ascending=[False, False]).head(20).reset_index(drop=True)
    merged.insert(0, '순위', range(1, len(merged) + 1))
    merged = merged[['순위', '물리대카테', 'Admin브랜드명', '고객수_당', '고객수증감',
                     '고객수증감률(%)', '거래액_당', '거래액전년비(%)', '고객수효과', '객단가효과']]
    merged.columns = ['순위', '카테고리', '브랜드', '고객수', '고객수증감',
                      '고객수증감률(%)', '거래액', '거래액전년비(%)', '고객수효과', '객단가효과']
    return merged


def score_and_select(df1, df_wow, df2, df3, top_n=30,
                     w1=0.30, w_wow=0.20, w2=0.30, w3=0.20):
    """
    [분석5] 가중치 종합점수 브랜드 선정
    각 분석의 순위 역수 기반 점수화 후 가중 합산
    """
    scores = {}
    keys = ['트렌드점수', '전주비점수', '볼륨점수', '신장률점수']

    def add(df, weight, score_key):
        if df is None or df.empty:
            return
        n = len(df)
        for i, (_, row) in enumerate(df.iterrows()):
            brand = row.get('브랜드', '')
            cat   = row.get('카테고리', '')
            score = (n - i) / n * weight * 100
            if brand not in scores:
                scores[brand] = dict({'브랜드': brand, '카테고리': cat},
                                     **{k: 0.0 for k in keys})
            scores[brand][score_key] += score

    add(df1,     w1,    '트렌드점수')
    add(df_wow,  w_wow, '전주비점수')
    add(df2,     w2,    '볼륨점수')
    add(df3,     w3,    '신장률점수')

    if not scores:
        return pd.DataFrame()

    result = pd.DataFrame(scores.values())
    result['종합점수'] = result[keys].sum(axis=1).round(1)
    for k in keys:
        result[k] = result[k].round(1)
    result = result.sort_values('종합점수', ascending=False).head(top_n).reset_index(drop=True)
    result.insert(0, '순위', range(1, len(result)+1))
    return result[['순위', '카테고리', '브랜드', '종합점수'] + keys]


def get_top_products(df_raw, week_map, target_week_code, selected_brands, top_n=10):
    """
    [분석5] 선택 브랜드의 상품코드 TOP 10 (고객수 기준)
    출력: 브랜드/자사입점/정상이월/상품코드/상품명/MD명/고객수/거래액
    """
    if df_raw is None or not selected_brands:
        return pd.DataFrame()

    df = tag_week(df_raw[df_raw['정산구분'] == '판매'], week_map)
    df = df[(df['_week'] == target_week_code) & df['Admin브랜드명'].isin(selected_brands)].copy()
    if df.empty:
        return pd.DataFrame()

    df['amt'] = get_vat_col(df)

    # 자사/입점 컬럼 처리 (제휴처구분1로 대체)
    입점col = '자사/입점' if '자사/입점' in df.columns else '제휴처구분1'

    grp_cols = ['Admin브랜드명', 입점col, '정상이월구분', '상품코드', '상품명', 'MD명']
    grp = df.groupby(grp_cols, as_index=False).agg(
        고객수=('고객번호', 'nunique'),
        거래액=('amt', 'sum')
    )

    parts = []
    for brand in selected_brands:
        b = grp[grp['Admin브랜드명'] == brand].nlargest(top_n, '고객수')
        parts.append(b)

    if not parts:
        return pd.DataFrame()

    final = pd.concat(parts, ignore_index=True)
    final = final.rename(columns={'Admin브랜드명': '브랜드', 입점col: '자사/입점', '정상이월구분': '정상/이월'})
    return final[['브랜드','자사/입점','정상/이월','상품코드','상품명','MD명','고객수','거래액']]


# ─────────────────────────────────────────────
# 세션 상태 초기화
# ─────────────────────────────────────────────
defaults = {
    'df_mall': None, 'df_aff_25': None, 'df_aff_26': None,
    'week_map_25': {}, 'week_map_26': {},
    'analysis_done': False,
    'df_prod': pd.DataFrame(),
    'df_wow': pd.DataFrame(),
    'base_prev_mall': '',
    'last_checks': '',
    'df1': pd.DataFrame(), 'df2': pd.DataFrame(),
    'df3': pd.DataFrame(), 'df4': pd.DataFrame(),
    'sel_week': '', 'sel_prev_mall_week': '',
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─────────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🏷️ 시크릿 큐레이션")
    st.markdown("---")
    st.markdown("### 📂 데이터 업로드")

    # ── ① 몰전체 CSV ──
    st.markdown("**① 2025년 몰 주차별 실적 (CSV)**")
    f_mall = st.file_uploader("mall", type=['csv'], label_visibility='collapsed', key='up_mall')
    if f_mall:
        if needs_parse('up_mall', f_mall):
            with st.spinner("로딩..."):
                df, err = load_mall_csv(f_mall.getvalue())
            if err:
                st.session_state.df_mall = None
                clear_parsed('up_mall')
                st.error(err)
            else:
                st.session_state.df_mall = df
                mark_parsed('up_mall', f_mall)
            del df
        if st.session_state.df_mall is not None:
            df = st.session_state.df_mall
            st.markdown('<span class="badge-ok">✓ 로드 완료</span>', unsafe_allow_html=True)
            st.caption(f"{len(df):,}행 · 브랜드 {df['ADMIN브랜드명'].nunique():,}개")
    else:
        st.markdown('<span class="badge-wait">⏳ 대기 중</span>', unsafe_allow_html=True)

    st.markdown("")

    # ── ② 2025 제휴 raw ──
    st.markdown("**② 2025년 분기 제휴 실적 (xlsx)**")
    f25 = st.file_uploader("aff25", type=['xlsx'], label_visibility='collapsed', key='up_aff25')
    if f25:
        if needs_parse('up_aff25', f25):
            with st.spinner("로딩..."):
                dr, dp, err = load_affiliate_excel(f25.getvalue(), '25')
            if err:
                st.session_state.df_aff_25 = None
                clear_parsed('up_aff25')
                st.error(err)
            else:
                st.session_state.df_aff_25 = dr
                if dp is not None:
                    st.session_state.week_map_25 = build_week_map(dp)
                mark_parsed('up_aff25', f25)
            del dr, dp
        if st.session_state.df_aff_25 is not None:
            st.markdown('<span class="badge-ok">✓ 로드 완료</span>', unsafe_allow_html=True)
            st.caption(f"{len(st.session_state.df_aff_25):,}행")
    else:
        st.markdown('<span class="badge-wait">⏳ 대기 중</span>', unsafe_allow_html=True)

    st.markdown("")

    # ── ③ 2026 제휴 raw ──
    st.markdown("**③ 2026년 당월 제휴 실적 (xlsx)**")
    f26 = st.file_uploader("aff26", type=['xlsx'], label_visibility='collapsed', key='up_aff26')
    if f26:
        if needs_parse('up_aff26', f26):
            with st.spinner("로딩..."):
                dr, dp, err = load_affiliate_excel(f26.getvalue(), '26')
            if err:
                st.session_state.df_aff_26 = None
                clear_parsed('up_aff26')
                st.error(err)
            else:
                st.session_state.df_aff_26 = dr
                if dp is not None:
                    st.session_state.week_map_26 = build_week_map(dp)
                mark_parsed('up_aff26', f26)
            del dr, dp
        if st.session_state.df_aff_26 is not None:
            dr = st.session_state.df_aff_26
            st.markdown('<span class="badge-ok">✓ 로드 완료</span>', unsafe_allow_html=True)
            st.caption(f"{len(dr):,}행 · {dr['정산일시일'].min()} ~ {dr['정산일시일'].max()}")
    else:
        st.markdown('<span class="badge-wait">⏳ 대기 중</span>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📅 분석 기간 설정")

    # 2026 주차코드 목록
    avail_26 = []
    if st.session_state.df_aff_26 is not None and st.session_state.week_map_26:
        wm = st.session_state.week_map_26
        dates = pd.to_datetime(st.session_state.df_aff_26['정산일시일']).dt.strftime('%Y-%m-%d').unique()
        avail_26 = sorted(set(wm.get(d,'') for d in dates if wm.get(d,'')))

    if avail_26:
        sel_week = st.selectbox("분석 주차 (2026)", avail_26, index=len(avail_26)-1)
    else:
        sel_week = st.text_input("분석 주차 (직접 입력, 예: 26_6_3)", value='26_6_3')

    # 전년 비교 주차: 몰전체 CSV 주차 목록에서 선택
    avail_mall_weeks = []
    if st.session_state.df_mall is not None:
        avail_mall_weeks = sorted(st.session_state.df_mall['결제_주차'].unique())

    # 게시할 주차의 트렌드를 미리 보려는 것이므로, 전년 '동주차'가 아니라
    # 전년 '다음 주차'를 기본값으로 잡는다. 예: 26_7_4 → 25_8_1
    prev_code = get_prev_week_code(sel_week)  # 26_7_4 → 25_7_4
    # 주차코드 → CSV 원본 표기. 코드로 계산한 주차를 데이터 표기로 되돌릴 때 쓴다.
    mall_week_by_code = {mall_week_to_code(x): x for x in avail_mall_weeks}
    mall_codes = set(mall_week_by_code)
    target_prev_code = (
        get_next_week_code(prev_code, mall_codes or None) if prev_code else None
    )
    # 데이터에 있는 표기를 우선 쓴다. code_to_mall_week 가 만드는 '07월' 형식이
    # CSV 표기('7월' 등)와 다르면 목록에서 못 찾아 엉뚱한 주차가 기본 선택된다.
    auto_prev_mall = ''
    if target_prev_code:
        auto_prev_mall = mall_week_by_code.get(
            target_prev_code, code_to_mall_week(target_prev_code)
        )

    if avail_mall_weeks:
        default_idx = avail_mall_weeks.index(auto_prev_mall) if auto_prev_mall in avail_mall_weeks else 0
        sel_prev_mall = st.selectbox(
            "전년 비교 주차 (몰전체 기준)",
            avail_mall_weeks, index=default_idx
        )
    else:
        sel_prev_mall = st.text_input("전년 비교 주차 (예: 2025년 06월 3주차)", value=auto_prev_mall)

    st.caption(f"분석: `{sel_week}` ↔ 전년: `{mall_week_to_code(sel_prev_mall) if sel_prev_mall else '-'}`")
    st.caption("전년 비교는 분석 주차의 **다음 주차**를 기본값으로 잡습니다 (예: 26_7_4 → 25_8_1).")

    st.markdown("---")
    with st.expander("⚙️ 가중치 설정"):
        w1    = st.slider("전년 트렌드",  0.0, 1.0, 0.30, 0.05)
        w_wow = st.slider("전주비 신장",  0.0, 1.0, 0.20, 0.05)
        w2    = st.slider("제휴 볼륨",    0.0, 1.0, 0.30, 0.05)
        w3    = st.slider("전년비 신장",  0.0, 1.0, 0.20, 0.05)
        tw = w1 + w_wow + w2 + w3
        if abs(tw - 1.0) > 0.02:
            st.warning(f"합계 {tw:.2f} (합계 1.0 권장)")

    st.markdown("---")
    st.markdown("### 📤 대시보드 내보내기")
    if st.session_state.analysis_done:
        st.download_button(
            "🌐 대시보드 HTML 다운로드",
            data=current_report_html().encode('utf-8'),
            file_name=f"시크릿_대시보드_{st.session_state.sel_week}_"
                      f"{datetime.now().strftime('%y%m%d')}.html",
            mime='text/html', use_container_width=True,
            help="표와 요약이 그대로 담긴 단일 HTML 파일입니다. "
                 "인터넷 없이 열리고 그대로 공유·인쇄할 수 있습니다.",
        )
    else:
        st.caption("분석 실행 후 활성화됩니다.")

    st.markdown("---")
    st.markdown("### 💾 Raw 데이터 내보내기")
    has_data = any(st.session_state[k] is not None for k in ['df_mall','df_aff_25','df_aff_26'])
    if has_data:
        json_bytes = dfs_to_json_bytes({
            'mall_2025': st.session_state.df_mall,
            'affiliate_2025': st.session_state.df_aff_25,
            'affiliate_2026': st.session_state.df_aff_26,
        })
        st.download_button(
            "📥 JSON으로 다운로드", json_bytes,
            file_name=f"lf_secret_raw_{datetime.now().strftime('%y%m%d')}.json",
            mime='application/json', use_container_width=True
        )
    else:
        st.caption("파일 업로드 후 활성화됩니다.")

    st.markdown("---")
    run_btn = st.button("🔍 분석 실행", use_container_width=True, type="primary")


# ─────────────────────────────────────────────
# 메인 영역
# ─────────────────────────────────────────────
st.markdown(f"""
<div class="header-banner">
  <div>
    <div class="header-title">🏷️ LF몰 시크릿 페이지 큐레이션 대시보드</div>
    <div class="header-sub">전년 트렌드 × 제휴 실적 × 신장률 → 게시 브랜드 & 상품 선정</div>
  </div>
  <div class="header-badge">내부 전용 · CONFIDENTIAL</div>
</div>
""", unsafe_allow_html=True)

# ── 분석 실행 ──
if run_btn:
    missing = []
    if st.session_state.df_mall is None:      missing.append("① 몰전체 CSV")
    if st.session_state.df_aff_26 is None:    missing.append("③ 2026 제휴 xlsx")
    if not st.session_state.week_map_26:      missing.append("피벗 시트 (2026 xlsx 내)")

    if missing:
        st.warning(f"필수 파일 누락: {', '.join(missing)}")
    else:
        with st.spinner("분석 중..."):
            df1 = analyze_mall_top20(st.session_state.df_mall, sel_prev_mall)

            # 전주비 신장: 선택된 전년 비교 주차(=게시 대상 주차) 직전 주차를 기준으로 삼는다.
            #
            # 기준 주차는 코드로 계산한 뒤 반드시 '데이터에 실제로 있는 표기'로
            # 되돌려야 한다. code_to_mall_week 는 월을 두 자리로 채우므로
            # ('2025년 07월 4주차') CSV 표기가 다르면 조용히 매칭에 실패한다.
            base_code = get_prev_adjacent_week_code(
                mall_week_to_code(sel_prev_mall), mall_codes or None
            )
            base_prev_mall = mall_week_by_code.get(
                base_code, code_to_mall_week(base_code) if base_code else ''
            )
            df_wow = analyze_mall_wow_growth(
                st.session_state.df_mall, base_prev_mall, sel_prev_mall
            )

            df2 = analyze_affiliate_top20_vol(
                st.session_state.df_aff_26, st.session_state.week_map_26, sel_week
            )
            df3 = analyze_affiliate_top20_growth(
                st.session_state.df_aff_26, st.session_state.df_aff_25,
                st.session_state.week_map_26, st.session_state.week_map_25,
                sel_week
            )
            df4 = score_and_select(df1, df_wow, df2, df3, top_n=30,
                                   w1=w1, w_wow=w_wow, w2=w2, w3=w3)

        st.session_state.df1 = df1
        st.session_state.df_wow = df_wow
        st.session_state.df2 = df2
        st.session_state.df3 = df3
        st.session_state.df4 = df4
        st.session_state.base_prev_mall = base_prev_mall
        st.session_state.sel_week = sel_week
        st.session_state.sel_prev_mall_week = sel_prev_mall
        st.session_state.analysis_done = True

        # 검증 로그
        checks = []
        checks.append(f"분석1 브랜드: {len(df1)}개 (전년 몰전체 {sel_prev_mall})")
        checks.append(f"분석2 브랜드: {len(df_wow)}개 (전주비 {base_prev_mall}→{sel_prev_mall})")
        if df_wow.empty:
            st.warning(
                f"전주비 신장 결과가 비어 있어 해당 가중치({w_wow:.0%})가 "
                f"최종 선정에 반영되지 않았습니다. "
                f"기준 주차 '{base_prev_mall}' 가 몰전체 데이터에 있는지 확인해주세요."
            )
        checks.append(f"분석3 브랜드: {len(df2)}개 (제휴 볼륨 {sel_week})")
        checks.append(f"분석4 브랜드: {len(df3)}개 (전년비 신장)")
        checks.append(f"최종 선정: {len(df4)}개")
        st.session_state.last_checks = " · ".join(checks)

        # 사이드바는 이 지점보다 먼저 그려지므로, 방금 끝난 분석 결과가
        # 사이드바(HTML 내보내기 버튼)에 반영되지 않는다. 한 번 다시 돌려
        # 사이드바가 완료 상태를 보게 한다.
        st.rerun()

if st.session_state.analysis_done and st.session_state.last_checks:
    st.success("분석 완료 ✓\n" + st.session_state.last_checks)

# ── KPI 요약 ──
if st.session_state.analysis_done:
    w = st.session_state.sel_week
    pw = st.session_state.sel_prev_mall_week
    n1,n2,n3,n4 = (len(st.session_state[f'df{i}']) for i in range(1,5))
    n_wow = len(st.session_state.df_wow)
    st.markdown(f"""
    <div class="kpi-row">
      <div class="kpi-card">
        <div class="kpi-label">전년 트렌드 TOP</div>
        <div class="kpi-value">{n1}</div>
        <div class="kpi-sub">몰전체 · {mall_week_to_code(pw)}</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-label">전주비 신장 TOP</div>
        <div class="kpi-value">{n_wow}</div>
        <div class="kpi-sub">몰전체 · 전주 대비</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-label">제휴 볼륨 TOP</div>
        <div class="kpi-value">{n2}</div>
        <div class="kpi-sub">제휴 · {w}</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-label">전년비 신장 TOP</div>
        <div class="kpi-value">{n3}</div>
        <div class="kpi-sub">전년 동주차 대비</div>
      </div>
      <div class="kpi-card" style="border-left-color:#D63E8A;">
        <div class="kpi-label">최종 선정</div>
        <div class="kpi-value" style="color:#D63E8A;">{n4}</div>
        <div class="kpi-sub">가중치 종합</div>
      </div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 탭
# ─────────────────────────────────────────────
tab1, tab_wow, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 ① 전년 트렌드 TOP 20",
    "🚀 ② 전년 전주비 신장 TOP 30",
    "💰 ③ 제휴 볼륨 TOP 20",
    "📈 ④ 전년비 신장 TOP 20",
    "🏷️ ⑤ 최종 선정 30개",
    "🛍️ ⑥ 상품코드 출력",
])

def render_table_tab(tab, num, title, desc, chip_cls, chip_txt, df_key,
                     fmt_cols=None, dl_sheet='Sheet1', dl_prefix=''):
    with tab:
        st.markdown(f"""
        <div class="sec-hdr">
          <span class="sec-num">{num}</span>
          <span class="sec-title">{title}</span>
          <span class="sec-desc">
            <span class="w-chip {chip_cls}">{chip_txt}</span>{desc}
          </span>
        </div>
        <div class="tbl-wrap">
        """, unsafe_allow_html=True)

        df = st.session_state[df_key]
        if st.session_state.analysis_done and not df.empty:
            df_show = df.copy()
            if fmt_cols:
                for col, fn in fmt_cols.items():
                    if col in df_show.columns:
                        df_show[col] = df_show[col].apply(fn)
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=550)
            st.download_button(
                "📥 엑셀 다운로드",
                data=df_to_excel_bytes(df, dl_sheet),
                file_name=f"{dl_prefix}_{datetime.now().strftime('%y%m%d')}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        elif st.session_state.analysis_done:
            st.markdown('<div class="warn-box">⚠️ 해당 주차 데이터 없음. 주차 선택 및 파일 확인 필요.</div>',
                        unsafe_allow_html=True)
        else:
            st.markdown('<div class="info-box">📌 사이드바에서 파일 업로드 후 <strong>분석 실행</strong>을 눌러주세요.</div>',
                        unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

render_table_tab(
    tab1, "1", "전년 몰전체 다음 주차 판매 TOP 20",
    "기준: 2025년 몰전체 CSV · 분석 주차의 다음 주차 · 거래액 내림차순",
    "trend", "트렌드 30%", "df1",
    fmt_cols={'거래액': fmt_amt, '고객수': fmt_num},
    dl_sheet='전년트렌드TOP20', dl_prefix='전년트렌드TOP20'
)

render_table_tab(
    tab_wow, "2", "전년 전주비 신장 TOP 30",
    "기준: 전년 분석주차 몰전체 거래액 TOP100 → 다음 주차 신장률 내림차순",
    "wow", "전주비 20%", "df_wow",
    fmt_cols={'거래액': fmt_amt, '고객수': fmt_num, '전주비(%)': fmt_pct},
    dl_sheet='전주비신장TOP30', dl_prefix='전주비신장TOP30'
)

render_table_tab(
    tab2, "3", "제휴 실적 분석주차 판매 TOP 20",
    "기준: 당월인증 Y/N 포함 · 판매건 · VAT 제외 · 거래액 내림차순",
    "vol", "볼륨 30%", "df2",
    fmt_cols={'거래액': fmt_amt, '객단가': fmt_amt, '고객수': fmt_num},
    dl_sheet='제휴볼륨TOP20', dl_prefix='제휴볼륨TOP20'
)

render_table_tab(
    tab3, "4", "전년 동주차 대비 고객수 신장 TOP 20",
    "기준: 당년·전년 고객수 5명↑ · 거래액 30만↑ & 전년비 플러스 · 고객수 절대 증감분 순 "
    "(고객수효과/객단가효과 = Δ거래액 요인분해)",
    "growth", "전년비 20%", "df3",
    fmt_cols={'고객수': fmt_num, '고객수증감': fmt_int_signed, '고객수증감률(%)': fmt_pct,
              '거래액': fmt_amt, '거래액전년비(%)': fmt_pct,
              '고객수효과': fmt_amt, '객단가효과': fmt_amt},
    dl_sheet='고객수신장TOP20', dl_prefix='고객수신장TOP20'
)

render_table_tab(
    tab4, "5", "최종 게시 브랜드 선정 (30개)",
    "가중치 종합점수 기준 · 사이드바 ⚙️에서 가중치 조정 가능",
    "trend", "종합", "df4",
    dl_sheet='최종선정30', dl_prefix='시크릿_선정브랜드30'
)

# ── 탭6: 상품코드 ──
with tab5:
    st.markdown("""
    <div class="sec-hdr">
      <span class="sec-num">6</span>
      <span class="sec-title">브랜드별 상품코드 TOP 10 출력</span>
      <span class="sec-desc">최종 선정 브랜드 중 상품코드 추출 대상을 선택하세요</span>
    </div>
    <div class="tbl-wrap">
    """, unsafe_allow_html=True)

    if st.session_state.analysis_done and not st.session_state.df4.empty:
        brand_opts = st.session_state.df4['브랜드'].tolist()
        selected = st.multiselect(
            "상품코드 출력할 브랜드 선택",
            options=brand_opts,
            default=brand_opts[:5],
            placeholder="브랜드 선택...",
        )

        if selected:
            with st.spinner("상품 데이터 추출 중..."):
                df_prod = get_top_products(
                    st.session_state.df_aff_26,
                    st.session_state.week_map_26,
                    st.session_state.sel_week,
                    selected, top_n=10
                )

            st.session_state.df_prod = df_prod  # HTML 내보내기에서 재사용
            if not df_prod.empty:
                df_show = df_prod.copy()
                df_show['거래액'] = df_show['거래액'].apply(fmt_amt)
                df_show['고객수'] = df_show['고객수'].apply(fmt_num)
                st.dataframe(df_show, use_container_width=True, hide_index=True, height=520)
                col_dl, col_info = st.columns([1, 3])
                with col_dl:
                    st.download_button(
                        "📥 상품코드 엑셀 다운로드",
                        data=df_to_excel_bytes(df_prod, '상품코드_TOP10'),
                        file_name=f"시크릿_상품코드_{datetime.now().strftime('%y%m%d')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                with col_info:
                    st.caption(f"선택 {len(selected)}개 브랜드 · 브랜드별 고객수 기준 TOP 10")
            else:
                st.markdown('<div class="warn-box">⚠️ 선택 브랜드의 해당 주차 상품 데이터 없음.</div>',
                            unsafe_allow_html=True)
        else:
            st.markdown('<div class="info-box">⬆️ 브랜드를 선택하면 상품코드가 출력됩니다.</div>',
                        unsafe_allow_html=True)
    else:
        st.markdown('<div class="info-box">📌 ④ 최종 선정 분석 완료 후 활성화됩니다.</div>',
                    unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

# ── 푸터 ──
st.markdown("---")
st.markdown(
    f"<div style='text-align:center;color:#9BA8B5;font-size:.72rem;'>"
    f"LF몰 제휴팀 내부 전용 · 무단 배포 금지 · "
    f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>",
    unsafe_allow_html=True
)
